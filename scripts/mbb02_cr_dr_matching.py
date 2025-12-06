import argparse
import csv
import json
import os
import re
from typing import List, Optional, Tuple, Dict, Any


def parse_mbb02_tran_desc(desc: str) -> Optional[Tuple[str, str, str]]:
    """
    Parse transaction description for MBB02 to extract type (CR/DR) and the MN/M/N number.

    Matches examples like:
      - "CR/CARD SALES MN 07700511 DATED 01072025" -> ("CR", "07700511")
      - "DR/CARD SALES M/N 8850927 DATED 01072025" -> ("DR", "8850927")

    Returns (tx_type, number, bank_date_iso) or None if not found/invalid.
    """
    if not desc:
        return None

    pattern = re.compile(
        r"(?i)\b(?P<type>CR|DR)\s*/\s*CARD\s+SALES\b.*?M\s*/?\s*N\s*(?P<number>\d{7,8})\b.*?\bDATED\b\s*(?P<date>\d{8})\b"
    )
    m = pattern.search(desc)
    if not m:
        return None
    tx_type = m.group("type").upper()
    number = m.group("number")
    date_raw = m.group("date")  # ddmmyyyy
    # Enforce length by type: CR -> 8, DR -> 7
    if tx_type == "CR" and len(number) != 8:
        return None
    if tx_type == "DR" and len(number) != 7:
        return None
    # Normalize date
    bank_date_iso = normalize_ddmmyyyy_to_iso(date_raw)
    if not bank_date_iso:
        return None
    return tx_type, number, bank_date_iso


def _digits(s: Any) -> str:
    return re.sub(r"\D", "", "" if s is None else str(s))


def read_c14_from_csv(path: str) -> Optional[str]:
    """Read cell C14 (row 14, column 3) from a CSV file as string (raw)."""
    try:
        with open(path, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader, start=1):
                if idx == 14:
                    # column C is index 2 (0-based)
                    if len(row) >= 3:
                        return row[2].strip()
                    return None
    except Exception:
        # Retry with default encoding
        with open(path, newline='') as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader, start=1):
                if idx == 14:
                    if len(row) >= 3:
                        return row[2].strip()
                    return None
    return None


def read_c14_from_excel(path: str) -> Optional[str]:
    """Read cell C14 from an Excel file via openpyxl (if available)."""
    try:
        import openpyxl  # type: ignore
    except Exception:
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        v = ws['C14'].value
        return ("" if v is None else str(v)).strip()
    except Exception:
        return None


def extract_reference_numbers_from_csv(path: str) -> List[str]:
    """Extract all values under 'Reference No' blocks until each 'Total Amount' row."""
    refs: List[str] = []
    # Load all rows
    try:
        with open(path, newline='', encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
    except Exception:
        with open(path, newline='') as f:
            rows = list(csv.reader(f))

    # Scan for header rows containing 'Reference No'
    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        # normalize row cells
        norm = [str(c).strip() if c is not None else "" for c in row]
        # find 'Reference No' column in this row
        try:
            col_idx = next(j for j, c in enumerate(norm) if c.lower() == 'reference no')
        except StopIteration:
            i += 1
            continue

        # collect values below until a 'Total Amount' row appears
        j = i + 1
        while j < n:
            r = rows[j]
            r_norm = [str(c).strip() if c is not None else "" for c in r]
            if any(isinstance(cell, str) and 'total amount' in cell.lower() for cell in r_norm):
                break
            # guard bounds
            if col_idx < len(r_norm):
                val = r_norm[col_idx]
                if val:
                    refs.append(val)
            j += 1
        i = j + 1  # skip past 'Total Amount' row

    # De-duplicate while preserving order
    seen = set()
    ordered: List[str] = []
    for r in refs:
        if r not in seen:
            seen.add(r)
            ordered.append(r)
    return ordered


def extract_reference_numbers_from_excel(path: str) -> List[str]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return []
    refs: List[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        r = 1
        while r <= max_row:
            # search for 'Reference No' in row r
            ref_col = None
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and val.strip().lower() == 'reference no':
                    ref_col = c
                    break
            if ref_col is None:
                r += 1
                continue
            # Collect below until 'Total Amount' row
            rr = r + 1
            while rr <= max_row:
                total_hit = False
                for cc in range(1, max_col + 1):
                    v = ws.cell(row=rr, column=cc).value
                    if isinstance(v, str) and 'total amount' in v.strip().lower():
                        total_hit = True
                        break
                if total_hit:
                    break
                v = ws.cell(row=rr, column=ref_col).value
                if v is not None and str(v).strip():
                    refs.append(str(v).strip())
                rr += 1
            r = rr + 1
    except Exception:
        return []
    # De-duplicate
    seen = set()
    out: List[str] = []
    for x in refs:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def get_last_row_m_o_from_csv(path: str) -> Tuple[Optional[str], Optional[str]]:
    try:
        with open(path, newline='', encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
    except Exception:
        with open(path, newline='') as f:
            rows = list(csv.reader(f))

    # Find last non-empty row
    last_idx = None
    for idx in range(len(rows) - 1, -1, -1):
        row = rows[idx]
        if any((str(c).strip() if c is not None else "") for c in row):
            last_idx = idx
            break
    if last_idx is None:
        return None, None
    row = rows[last_idx]
    m_val = row[12].strip() if len(row) > 12 and row[12] is not None else None  # Column M (0-based 12)
    o_val = row[14].strip() if len(row) > 14 and row[14] is not None else None  # Column O (0-based 14)
    m_val = m_val if m_val != '' else None
    o_val = o_val if o_val != '' else None
    return m_val, o_val


def get_last_row_m_o_from_excel(path: str) -> Tuple[Optional[str], Optional[str]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return None, None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        # Find last non-empty row by scanning upward
        last = ws.max_row or 0
        def row_has_data(r: int) -> bool:
            for c in range(1, (ws.max_column or 0) + 1):
                v = ws.cell(row=r, column=c).value
                if v not in (None, ''):
                    return True
            return False
        while last > 0 and not row_has_data(last):
            last -= 1
        if last <= 0:
            return None, None
        m_val = ws.cell(row=last, column=13).value  # M
        o_val = ws.cell(row=last, column=15).value  # O
        m_str = ("" if m_val is None else str(m_val)).strip() or None
        o_str = ("" if o_val is None else str(o_val)).strip() or None
        return m_str, o_str
    except Exception:
        return None, None


def to_number_or_str(v: Optional[str]) -> Any:
    if v is None:
        return None
    s = str(v).replace(',', '').strip()
    try:
        if s == '':
            return None
        return float(s)
    except Exception:
        return s


def normalize_ddmmyyyy_to_iso(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    s = re.sub(r"\D", "", str(s))
    if len(s) != 8:
        return None
    dd, mm, yyyy = s[:2], s[2:4], s[4:]
    try:
        d = int(dd); m = int(mm); y = int(yyyy)
        from datetime import date
        return date(y, m, d).isoformat()
    except Exception:
        return None


def normalize_generic_date_to_iso(v: Any) -> Optional[str]:
    """Normalize various CSV/Excel date representations to ISO YYYY-MM-DD."""
    from datetime import date, datetime
    if v is None:
        return None
    # If already datetime/date
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # 8 digits ddmmyyyy
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        iso = normalize_ddmmyyyy_to_iso(digits)
        if iso:
            return iso
    # d/m/yyyy or dd/mm/yyyy (Malaysia typical)
    if "/" in s:
        parts = [p.strip() for p in s.split("/")]
        if len(parts) == 3:
            try:
                d = int(parts[0]); m = int(parts[1]); y = int(parts[2])
                from datetime import date
                return date(y, m, d).isoformat()
            except Exception:
                pass
    # yyyy-mm-dd
    try:
        from datetime import datetime as _dt
        return _dt.strptime(s, "%Y-%m-%d").date().isoformat()
    except Exception:
        pass
    # yyyy/m/d
    try:
        from datetime import datetime as _dt
        return _dt.strptime(s, "%Y/%m/%d").date().isoformat()
    except Exception:
        pass
    return None


def read_b4_from_csv(path: str) -> Optional[str]:
    try:
        with open(path, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader, start=1):
                if idx == 4:
                    # B4 is index 1
                    if len(row) >= 2:
                        return row[1].strip()
                    return None
    except Exception:
        with open(path, newline='') as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader, start=1):
                if idx == 4:
                    if len(row) >= 2:
                        return row[1].strip()
                    return None
    return None


def read_b4_from_excel(path: str) -> Optional[str]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        v = ws['B4'].value
        return ("" if v is None else str(v)).strip()
    except Exception:
        return None


def _process_for_desc_core(
    bank_name: str,
    tran_desc: str,
    input_folder: str,
    output_folder: Optional[str] = None,
    write_when_empty: bool = False,
    persist_output: bool = True,
) -> Dict[str, Any]:
    """
    Shared core processing for a single bank transaction description.

    When `persist_output` is False, the JSON file will not be written even if
    matches are found. The original `process_for_desc` wrapper keeps legacy
    behaviour intact for the CLI entrypoint.
    """
    if bank_name.strip().lower() != 'mbb02':
        return {
            'bank_name': bank_name,
            'skipped': True,
            'reason': 'Bank name not Mbb02',
        }

    parsed = parse_mbb02_tran_desc(tran_desc)
    if not parsed:
        return {
            'bank_name': bank_name,
            'tran_desc': tran_desc,
            'matched': False,
            'error': 'Tran. Desc does not match MBB02 CR/DR pattern',
        }
    tx_type, bank_number, bank_date_iso = parsed
    trailing_len = 8 if tx_type == 'CR' else 7

    # enumerate candidate files
    files: List[str] = []
    for root, _, filenames in os.walk(input_folder):
        for fn in filenames:
            if fn.lower().endswith(('.csv', '.xlsx', '.xls')):
                files.append(os.path.join(root, fn))

    matched_files: List[Dict[str, Any]] = []
    for fp in files:
        ext = os.path.splitext(fp)[1].lower()
        c14_raw: Optional[str] = None
        b4_raw: Optional[str] = None
        if ext == '.csv':
            c14_raw = read_c14_from_csv(fp)
            b4_raw = read_b4_from_csv(fp)
        else:
            c14_raw = read_c14_from_excel(fp)
            b4_raw = read_b4_from_excel(fp)

        if not c14_raw:
            continue
        digits = _digits(c14_raw)
        if not digits or len(digits) < trailing_len:
            continue
        trailing = digits[-trailing_len:]
        if trailing != bank_number:
            continue
        # Date check: B4 must match the bank date (normalized)
        file_date_iso = normalize_generic_date_to_iso(b4_raw)
        if not file_date_iso or file_date_iso != bank_date_iso:
            continue

        # Extract reference numbers
        if ext == '.csv':
            references = extract_reference_numbers_from_csv(fp)
            m_val, o_val = get_last_row_m_o_from_csv(fp)
        else:
            references = extract_reference_numbers_from_excel(fp)
            m_val, o_val = get_last_row_m_o_from_excel(fp)

        matched_files.append({
            'file_path': fp,
            'c14_raw': c14_raw,
            'matched_trailing': trailing,
            'b4_raw': b4_raw,
            'file_date_iso': file_date_iso,
            'reference_numbers': references,
            'disc_amt': to_number_or_str(m_val),
            'net_amount': to_number_or_str(o_val),
        })

    result: Dict[str, Any] = {
        'bank_name': 'Mbb02',
        'tran_desc': tran_desc,
        'transaction_type': tx_type,
        'bank_number': bank_number,
        'bank_date_iso': bank_date_iso,
        'matched_files': matched_files,
    }

    # Only write output when at least one file matched, unless explicitly allowed
    if persist_output and output_folder and (matched_files or write_when_empty):
        os.makedirs(output_folder, exist_ok=True)
        safe_num = re.sub(r"[^0-9]", "", bank_number)
        out_name = f"mbb02_match_{tx_type.lower()}_{safe_num}_{bank_date_iso}.json"
        out_path = os.path.join(output_folder, out_name)
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        result['output_file'] = out_path
    return result


def process_for_desc(
    bank_name: str,
    tran_desc: str,
    input_folder: str,
    output_folder: str,
    write_when_empty: bool = False,
) -> Dict[str, Any]:
    """
    Core processing for a single bank transaction description.
    """
    return _process_for_desc_core(
        bank_name=bank_name,
        tran_desc=tran_desc,
        input_folder=input_folder,
        output_folder=output_folder,
        write_when_empty=write_when_empty,
        persist_output=True,
    )


def process_for_desc_in_memory(
    bank_name: str,
    tran_desc: str,
    input_folder: str,
) -> Dict[str, Any]:
    """
    Variant of `process_for_desc` that avoids writing JSON to disk.
    Returns the same dictionary structure but keeps the intermediate data in
    memory for consumers that wish to post-process the matched files.
    """
    return _process_for_desc_core(
        bank_name=bank_name,
        tran_desc=tran_desc,
        input_folder=input_folder,
        output_folder=None,
        write_when_empty=False,
        persist_output=False,
    )


def main():
    parser = argparse.ArgumentParser(description='MBB02 CR/DR Matching Script (requires bank JSON)')
    parser.add_argument('--bank-name', required=True, help='Bank name (run only if Mbb02)')
    parser.add_argument('--bank-json', required=True, help='Path to bank transactions JSON file (list of records)')
    parser.add_argument('--desc-field', required=False, default='Tran. Desc', help='Bank JSON field for description (default: Tran. Desc)')
    parser.add_argument('--input-folder', default='CreditCardTransaction', help='Folder containing credit card files')
    parser.add_argument('--output-folder', default='CreditCardTransactionGroupingOutput', help='Output folder for JSON result')
    parser.add_argument('--write-empty', dest='write_empty', action='store_true', help='Write JSON even when no matched file found (default: False)')
    parser.set_defaults(write_empty=False)
    args = parser.parse_args()

    bank_path = args.bank_json
    if not os.path.isfile(bank_path):
        raise SystemExit(f"Bank JSON not found: {bank_path}")
    if os.path.splitext(bank_path)[1].lower() != '.json':
        raise SystemExit("Bank input must be a .json file (Excel not supported)")
    outputs: List[Dict[str, Any]] = []
    seen = set()

    try:
        with open(bank_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        with open(bank_path, 'r') as f:
            data = json.load(f)
    if not isinstance(data, list):
        raise SystemExit("Bank JSON must be a list of records")
    for row in data:
        desc = str((row or {}).get(args.desc_field, '') or '')
        parsed = parse_mbb02_tran_desc(desc)
        if not parsed:
            continue
        tx_type, number, iso_date = parsed
        key = (tx_type, number, iso_date)
        if key in seen:
            continue
        seen.add(key)
        res = process_for_desc(
            bank_name=args.bank_name,
            tran_desc=desc,
            input_folder=args.input_folder,
            output_folder=args.output_folder,
            write_when_empty=bool(args.write_empty),
        )
        if res.get('matched_files'):
            outputs.append(res)

    print(json.dumps({
        'bank_name': args.bank_name,
        'bank_json': bank_path,
        'results_count': len(outputs),
        'results': outputs,
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
